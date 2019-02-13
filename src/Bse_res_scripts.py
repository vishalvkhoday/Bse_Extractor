'''
Created on Jan 10, 2019

@author: vkhoday
'''
import unittest
from Object_repository import driver
from selenium.webdriver.common.keys import Keys
import Object_repository as Obj_R
from  openpyxl import load_workbook
import time
import os

Bse_file = "C:/Users/vkhoday/git/Selenium_NSE_Algo/Additonal_Utility/Bse_Res_Script10Jan2019.xlsx"


class Test(unittest.TestCase):

    def setUp(self):
        driver.get("https://www.bseindia.com")
        driver.maximize_window()
        driver.implicitly_wait(5)
        pass

    def tearDown(self):
        pass

    def testName(self):
        
        Wb = load_workbook(Bse_file)
        WsScripts = Wb['Sheet1']
        Ws_Result = Wb['Results']
        rowCnt = WsScripts.max_row
        if rowCnt ==1:
            rowCnt=2
        
        print (rowCnt)
        for i in range(2, rowCnt):
            str_script_col = str('A') + str(i)
            Col_INIE = 'B' + str(i)
            col_st = str('N') + str(i)
            script_name = WsScripts[str_script_col].value
            INIE = WsScripts[Col_INIE].value
            exe_st = WsScripts[col_st].value            
            print ("{}) {}  {}".format(i,script_name,exe_st))
            if exe_st == 'Yes':
                
                Bse_sctTxt = driver.find_element_by_xpath(Obj_R.xpath_scr_txt)                
                Bse_sctTxt.send_keys(INIE)
                time.sleep(2)
                driver.set_page_load_timeout(2)
                try:
                    #Bse_sctTxt.send_keys(Keys.ENTER)
                    os.system("C:/Users/vkhoday/git/Selenium_NSE_Algo/Additonal_Utility/Enter.vbs")
                except:
                    print('Enter key occured')
                    None
                time.sleep(3)
                try:                              
                    driver.find_element_by_xpath(Obj_R.xpath_res_bt).click()
                    time.sleep(3)
                    href= driver.find_element_by_xpath('//*[@id="res"]/div/div[1]/table/tbody[7]/tr/td[2]/a').get_attribute('href')
                    print(href)
                except:
                    #href = driver.find_element_by_xpath(Obj_R.xpath_href_lk).get_attribute('pathname')
                    #driver.get(href)
                    None
                
                t_Tbl_details = driver.find_element_by_xpath(Obj_R.xpath_res_tbl).get_attribute('innerText')
                #t_Tbl_details = t_Tbl_details.replace('(in Cr.)', '')
                t_Tbl_details = t_Tbl_details.replace("Income Statement", "")
                spt_Tbl_details = t_Tbl_details.splitlines()
                Res_cnt = Ws_Result.max_row
                for rows in spt_Tbl_details:
                    print (rows)
                    spt_row = str(rows).split("\t")
                    if len(spt_row) > 1:
                        try:
                            if str(spt_row[2]).isalpha():
                                break
                            else:
                                Ws_Result['A'+ str(Res_cnt)]=script_name
                                Ws_Result['B'+ str(Res_cnt)]=spt_row[0]
                                Ws_Result['C'+ str(Res_cnt)]=spt_row[1]
                                Ws_Result['D'+ str(Res_cnt)]=spt_row[2]
                                Ws_Result['E'+ str(Res_cnt)]=spt_row[3]
                                Ws_Result['F'+ str(Res_cnt)]=spt_row[4]
                                Ws_Result['G'+ str(Res_cnt)]=spt_row[5]
                                Res_cnt+=1
                            
                        except:
                            None
                    temp_row = 'N'+ str(i)
                    WsScripts[temp_row]="No"
                    Wb.save(Bse_file)
                    continue
                WsScripts[temp_row]="No"
                Wb.save(Bse_file)
        pass


if __name__ == "__main__":
    # import sys;sys.argv = ['', 'Test.testName']
    unittest.main()
