'''
Created on Oct 29, 2019

@author: DELL
'''

import unittest
from Object_repository import driver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.actions.mouse_button import MouseButton
from selenium.webdriver.support.wait import WebDriverWait
import Object_repository as Obj_R
from  openpyxl import load_workbook
import time
import os
from selenium.webdriver.support import expected_conditions as EC
from allure_pytest.plugin import allure as alRep


# Bse_file = "C:/Users/DELL/git/Selenium_NSE_Algo/Additonal_Utility/Bse_Res_Script_14Apr2019.xlsx"
Bse_file = "C:/Users/DELL/git/Selenium_NSE_Algo/Additonal_Utility/BSE_Res_Scriptyear 19-20.xlsx"
# Bse_file = "C:/Users/DELL/git/Selenium_NSE_Algo/Additonal_Utility/Bse_Res_Script_14Sep2019.xlsx"
driver = Obj_R.driver
Wb = load_workbook(Bse_file)
WsScripts = Wb['Sheet1']
Ws_Result = Wb['Results']

class Test(unittest.TestCase):
    def clear_Temp (self):
            
        try:
            os.system('del /f/s/q %temp%\*')
        except AssertionError as ae:
            print('Unable to delete temp files')
    
    def setUp(self):
        driver.get("https://www.bseindia.com")
        alRep.testcase("https://www.bseindia.com", name=None)
        driver.maximize_window()
        driver.implicitly_wait(5)
        pass

    def tearDown(self):
        driver.quit()
        Wb.close()
        pass
    
    def testName(self):
        win_hand =driver.window_handles
        print ('{}'.format(win_hand))       
        rowCnt = WsScripts.max_row
        if rowCnt ==1:
            rowCnt=2
        
        print (rowCnt)
        icount =0
        for i in range(2, rowCnt):
            str_script_col = str('A') + str(i)
            Col_INIE = 'B' + str(i)
            col_st = str('N') + str(i)
            script_name = WsScripts[str_script_col].value
            INIE = WsScripts[Col_INIE].value
            exe_st = WsScripts[col_st].value            
            print ("{}) {}  {}".format(i,script_name,exe_st))
            
            if exe_st == 'Yes':
                try:
                    Bse_sctTxt = driver.find_element_by_xpath(Obj_R.xpath_scr_txt)                
                    Bse_sctTxt.send_keys(INIE)
                    time.sleep(3)
                    
                    if icount> 25:
                        Test.clear_Temp(self)
                        icount=0
                    else:
                        icount = icount+1
                    
#                     driver.switch_to_default_content
                    Bse_sctTxt.send_keys(Keys.ENTER)
                    time.sleep(1)
                    driver.find_element_by_xpath(Obj_R.xpath_scr_txt).clear()
                    secID=""
                    sec_code = ""
                    sec_ISIN = ""
                    scr_info =None
                    scr_info =driver.find_element_by_xpath(Obj_R.xpath_script_ID_INIE).get_attribute('innerText')
                    scr_info = scr_info.replace("(","")
                    scr_info = scr_info.replace(")","")
                    temp_scrId = str(scr_info).split("|")
                    if str(scr_info).find(INIE)==-1:
#                     if str(temp_scrId[2]).strip() != str(INIE).strip():
#                         driver.switch_to_default_content
                        driver.find_element_by_xpath(Obj_R.xpath_scr_txt).click()
                        driver.find_element_by_xpath(Obj_R.xpath_scr_txt).send_keys(script_name)
#                         Bse_sctTxt.send_keys(script_name)
                        driver.find_element_by_xpath(Obj_R.xpath_scr_txt).send_keys(Keys.ENTER)
                        driver.find_element_by_xpath(Obj_R.xpath_scr_txt).clear()
                        time.sleep(2)
                        
                    driver.find_element_by_xpath(Obj_R.xpath_tblHeader).location_once_scrolled_into_view
#                     tblHeader =driver.find_element_by_xpath(Obj_R.xpath_tblHeader).get_attribute('innerText')
#                     if tblHeader.find('Jun-19') == -1:
#                         continue
#                     else:
#                     ActionChains(driver).click('//*[@id="tabres"]')
#                     driver.find_element_by_xpath('//*[@id="res"]/div/div[1]/table/thead').location_once_scrolled_into_view
#                     time.sleep(2)
                    elm_res_bt = driver.find_element_by_xpath(Obj_R.xpath_res_bt)
                    time.sleep(1)
                    ActionChains(driver).move_to_element(elm_res_bt).click(elm_res_bt).perform()
                    ActionChains(driver).key_down(Keys.TAB).perform()
                    ActionChains(driver).key_down(Keys.ENTER).perform()
                    time.sleep(2)
                
                except:
                    print("Error occured")
                    pass

                try:
#                     driver.find_element_by_xpath(Obj_R.xpath_res_tbl).location_once_scrolled_into_view
                    driver.find_element_by_xpath('(//*[@id="aanualtrd"])[1]').location_once_scrolled_into_view
                    x_yrTab =driver.find_element_by_xpath('(//*[@id="aanualtrd"])[1]')
                    ActionChains(driver).move_to_element(x_yrTab).click(x_yrTab).perform()
                    t_Tbl_details = driver.find_element_by_xpath('//*[@id="ann"]').get_attribute('innerText')
                    t_Tbl_details = t_Tbl_details.replace("Income Statement", "").replace("%", "")
                    time.sleep(2)
#               driver.find_element_by_xpath('//*[@id="tabres"]').click()
                    spt_Tbl_details = t_Tbl_details.splitlines()
                    Res_cnt = Ws_Result.max_row
                    if Res_cnt ==1:
                        Res_cnt =2
                    
                    secID=""
                    sec_code = ""
                    sec_ISIN = ""
                    scr_info =None
                    scr_info =driver.find_element_by_xpath(Obj_R.xpath_script_ID_INIE).get_attribute('innerText')
                    scr_info = scr_info.replace("(","")
                    scr_info = scr_info.replace(")","")
                    temp_scrId = str(scr_info).split("|")
                    temp_scrId = temp_scrId
                    secID = temp_scrId[0]
                    sec_code =temp_scrId[1]
                    sec_ISIN = temp_scrId[2]
                    print(sec_code,secID)
                    if str(sec_ISIN.strip()) != str(INIE).strip():
                        continue
                except:
                    continue
                
                try:
                    for rows in range(0,len(spt_Tbl_details)-1):
                        print (spt_Tbl_details[rows])
                        spt_row = str(spt_Tbl_details[rows]).split("\t")
                        if len(spt_row) > 1:
                            try:
                                if str(spt_row[2]).isalpha():
                                    break
                                else:
                                    Ws_Result['A'+ str(Res_cnt)]=script_name
                                    Ws_Result['B'+ str(Res_cnt)]=sec_ISIN
                                    Ws_Result['C'+ str(Res_cnt)]=str(spt_row[0])
                                    Ws_Result['D'+ str(Res_cnt)]=str(spt_row[1])
                                    Ws_Result['E'+ str(Res_cnt)]=str(spt_row[2])
                                    Ws_Result['F'+ str(Res_cnt)]=str(spt_row[3])
                                    Ws_Result['G'+ str(Res_cnt)]=str(spt_row[4])
                                    Ws_Result['H'+ str(Res_cnt)]=str(spt_row[5])
                                    Res_cnt+=1
                                                        
                            except:
                                pass
#                                 temp_row = 'N'+ str(i)
#                                 WsScripts[temp_row]="No"
#                                 Wb.save(Bse_file)
                        
                    temp_row = 'N'+ str(i)
                    WsScripts[temp_row]="No"
                    Wb.save(Bse_file)
                    continue

                except:
                    temp_row = 'N'+ str(i)
                    WsScripts[temp_row]="No"
                    Wb.save(Bse_file)
                    continue
#                 finally:
#                     driver.quit()

if __name__ == "__main__":
    # import sys;sys.argv = ['', 'Test.testName']
    unittest.main()
    